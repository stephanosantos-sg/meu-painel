"""Parse 'Controle Financeiro - Stephano.xlsx' (2023+) and emit JSON for Orbita."""
import openpyxl, json, re, sys
from datetime import date, datetime

XLSX = "/Users/stephano/Downloads/Controle Financeiro - Stephano.xlsx"
OUT = "/Users/stephano/Downloads/orbita-v2/financas-import.json"

MONTHS_PT = ['JAN','FEV','MAR','ABR','MAI','JUN','JUL','AGO','SET','OUT','NOV','DEZ']

ACCOUNTS = {
    # exact normalized → accountId
    'pix sheida': 'pix-sheida',
    'pix babajam': 'pix-babajam',
    'pix lica': 'pix-lica',
    'pix stephano': 'pix-stephano',
    'pix': 'pix-stephano',
    'nubank sheida': 'nubank-sheida',
    'nubank': 'nubank-sheida',
    'visa sheida': 'visa-sheida',
    'visa': 'visa-sheida',
    'visa platinum': 'visa-sheida',
    'itau stephano': 'itau-stephano',
    'itau': 'itau-stephano',
    'itaú': 'itau-stephano',
    'itau credito': 'itau-stephano',
    'boleto': 'boleto',
    'dinheiro': 'dinheiro',
    'pessoal': 'itau-stephano',  # legacy default for Stephano
    'manu': 'pix-stephano',
    'ambos': 'itau-stephano',
    'visionary': 'itau-stephano',
}

# Word-based hints inside description when MEIO is generic
DESC_HINTS = [
    ('nubank', 'nubank-sheida'),
    ('visa', 'visa-sheida'),
    ('itau', 'itau-stephano'),
    ('itaú', 'itau-stephano'),
    ('boleto', 'boleto'),
    ('infinite', 'itau-stephano'),
    ('platinum', 'visa-sheida'),
]

CATEGORIES = {
    'saude': 'saude', 'saúde': 'saude',
    'moradia': 'moradia',
    'aluguel': 'moradia',
    'comida': 'comida',
    'restaurante': 'comida', 'restaurantes': 'comida',
    'cafés e restaurantes': 'comida', 'cafés': 'comida', 'café': 'comida',
    'alimentação': 'comida', 'alimentacao': 'comida',
    'marmita': 'comida', 'marmitas': 'comida',
    'delivery': 'delivery',
    'mercado': 'mercado', 'supermercado': 'mercado',
    'serviços': 'servicos', 'servicos': 'servicos',
    'assinaturas': 'servicos', 'streaming': 'servicos',
    'compras': 'compras',
    'lazer': 'lazer',
    'presentes': 'presentes',
    'dívidas': 'dividas', 'dividas': 'dividas',
    'cartão de crédito': 'cartao', 'cartao de credito': 'cartao',
    'cartão crédito': 'cartao',
    'investimento': 'investimento',
    'investimentos': 'investimento',
    'impostos e taxas': 'impostos', 'impostos': 'impostos',
    'nina': 'nina', 'catarina': 'nina',
    'transporte': 'transporte', 'uber': 'transporte', 'gasolina': 'transporte', 'combustível': 'transporte',
    'saídas': 'lazer',
    'necessidades básicas': 'compras',
}

def norm(s):
    if s is None: return ''
    return str(s).strip().lower()

def to_num(v):
    if v is None: return None
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip().replace('R$', '').replace(' ', '')
    s = s.replace('.', '').replace(',', '.') if s.count(',') == 1 and s.count('.') >= 1 else s.replace(',', '.')
    try: return float(s)
    except: return None

def map_account(meio, desc):
    n = norm(meio)
    if n in ACCOUNTS: return ACCOUNTS[n]
    # Try description hints
    dn = norm(desc)
    for hint, aid in DESC_HINTS:
        if hint in dn: return aid
    # Generic fallbacks for "PESSOAL"
    if n == 'pessoal':
        for hint, aid in DESC_HINTS:
            if hint in dn: return aid
        return 'itau-stephano'
    return 'itau-stephano'  # default

def map_category(cat):
    n = norm(cat)
    return CATEGORIES.get(n, 'compras' if n else 'compras')

def parse_installment(parcela):
    """Parse '16/21' → (16, 21). Returns None for dates or empty."""
    if parcela is None: return None
    s = str(parcela).strip()
    m = re.match(r'^(\d+)\s*/\s*(\d+)$', s)
    if m: return (int(m.group(1)), int(m.group(2)))
    return None

def parse_day(d, default=1):
    if d is None: return default
    if isinstance(d, (int, float)):
        i = int(d)
        if 1 <= i <= 31: return i
    if isinstance(d, datetime):
        return d.day
    s = str(d).strip()
    m = re.match(r'^\d+', s)
    if m:
        i = int(m.group(0))
        if 1 <= i <= 31: return i
    return default

def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)

    # Find sheets ≥ 2023
    target = []
    for name in wb.sheetnames:
        m = re.match(r'^Finanças (\w{3})(\d{2})$', name)
        if m:
            mon, yr = m.groups()
            if mon in MONTHS_PT:
                year = 2000 + int(yr)
                if year >= 2023:
                    target.append((year, MONTHS_PT.index(mon)+1, name))
    target.sort()
    print(f"Processing {len(target)} sheets from 2023+", file=sys.stderr)

    transactions = []
    income_by_month = {}
    debug_unmapped_cats = {}
    debug_unmapped_accs = {}

    for year, month, sheet_name in target:
        ws = wb[sheet_name]
        ym = f"{year}-{month:02d}"
        rows = list(ws.iter_rows(values_only=True))

        # Income — row 4 column B
        if len(rows) >= 4:
            inc = to_num(rows[3][1]) if len(rows[3]) > 1 else None
            if inc and inc > 0:
                income_by_month[ym] = inc

        # Find header row with GASTOS
        header_row = None
        for i in range(min(15, len(rows))):
            if any(norm(c) == 'gastos' for c in rows[i] if c):
                header_row = i
                break
        if header_row is None: continue

        # Parse transactions starting after header_row + 1
        for i in range(header_row + 1, len(rows)):
            row = rows[i]
            if not row or all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
                continue
            # Cols: A=✓, B=GASTOS, C=TIPO/MEIO, D=CATEGORIA, E=STATUS, F=PARCELA, G=SAÍDA, H=(blank/atualizado), I=DIA
            desc = (row[1] if len(row) > 1 else None)
            meio = (row[2] if len(row) > 2 else None)
            cat = (row[3] if len(row) > 3 else None)
            status = norm(row[4] if len(row) > 4 else '')
            parcela = (row[5] if len(row) > 5 else None)
            valor = to_num(row[6] if len(row) > 6 else None)
            dia_col = (row[8] if len(row) > 8 else None) or (row[7] if len(row) > 7 else None)

            # Skip empty/header/footer-ish rows
            if not desc or not isinstance(desc, str): continue
            if not valor or valor <= 0: continue
            d = desc.strip()
            if d.upper() in ('GASTOS', 'TOTAL', 'SOBRAS:', 'SUBTOTAL'): continue
            if d.startswith('SOBRAS'): continue

            # Track unmapped for debug
            cat_n = norm(cat)
            if cat_n and cat_n not in CATEGORIES:
                debug_unmapped_cats[cat_n] = debug_unmapped_cats.get(cat_n, 0) + 1
            meio_n = norm(meio)
            if meio_n and meio_n not in ACCOUNTS:
                debug_unmapped_accs[meio_n] = debug_unmapped_accs.get(meio_n, 0) + 1

            account_id = map_account(meio, desc)
            category_id = map_category(cat)
            installment = parse_installment(parcela)
            day = parse_day(dia_col, default=5)
            day = min(day, 28)

            tx = {
                'id': f'imp-{ym}-{i}',
                'description': d.strip(),
                'value': round(valor, 2),
                'date': f"{ym}-{day:02d}",
                'accountId': account_id,
                'categoryId': category_id,
                'status': 'paid' if 'resolv' in status else ('paid' if not status else 'pending'),
                'imported': True,
            }
            if installment:
                cur, total = installment
                tx['installment'] = {'current': cur, 'total': total}
            transactions.append(tx)

    # Investimentos sheet
    investments = []
    contributions = []
    if 'INVESTIMENTOS' in wb.sheetnames:
        ws = wb['INVESTIMENTOS']
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > 50: break
            if i < 5: continue  # skip header-ish
            # Try to detect rows with text+number
            if not row: continue
            name_cell = next((c for c in row if isinstance(c, str) and len(c.strip()) > 2), None)
            num_cell = next((c for c in row if isinstance(c, (int, float)) and c > 0), None)
            if name_cell and num_cell:
                investments.append({'rawName': name_cell.strip(), 'value': num_cell})

    out = {
        'transactions': transactions,
        'income_by_month': income_by_month,
        'investments_raw': investments,
        '_debug': {
            'sheets_processed': len(target),
            'unmapped_categories': dict(sorted(debug_unmapped_cats.items(), key=lambda x: -x[1])[:20]),
            'unmapped_accounts': dict(sorted(debug_unmapped_accs.items(), key=lambda x: -x[1])[:20]),
        }
    }

    print(f"\nGenerated {len(transactions)} transactions across {len(target)} months", file=sys.stderr)
    print(f"Income tracked for {len(income_by_month)} months", file=sys.stderr)
    print(f"Top unmapped categories: {list(out['_debug']['unmapped_categories'].items())[:10]}", file=sys.stderr)
    print(f"Top unmapped accounts: {list(out['_debug']['unmapped_accounts'].items())[:10]}", file=sys.stderr)

    with open(OUT, 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, indent=None)
    print(f"\nSaved to {OUT}", file=sys.stderr)

    # Stats
    by_year = {}
    for t in transactions:
        y = t['date'][:4]
        by_year[y] = by_year.get(y, 0) + 1
    print("\nBy year:", file=sys.stderr)
    for y, n in sorted(by_year.items()):
        total = sum(t['value'] for t in transactions if t['date'].startswith(y))
        print(f"  {y}: {n} txs, R$ {total:,.2f}", file=sys.stderr)

if __name__ == '__main__':
    main()
